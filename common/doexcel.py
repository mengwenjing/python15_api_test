from openpyxl import load_workbook

class Case:
    """
    测试数据类，每个测试数据，实际上就是它的一个实例
    """

    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None

class DoExcel:

    def __init__(self,filename,sheetname):
        try:
            self.filename = filename
            self.workbook = load_workbook(filename)
            self.sheet = self.workbook[sheetname]
        except Exception as e:
            print('找不到该文件，错误是：{}'.format(e))

    def read_data(self):
        big_data = []#定义一个空列表，用来存放所有的测试数据
        for i in range(2,self.sheet.max_row+1):#指定读取excel中1~6列的数据，将每一列的数据赋值给对象属性，最后全部添加到big_data列表中
            data = Case()
            data.case_id = self.sheet.cell(row=i,column=1).value
            data.title = self.sheet.cell(row=i, column=2).value
            data.url = self.sheet.cell(row=i, column=3).value
            data.data = self.sheet.cell(row=i, column=4).value
            data.method = self.sheet.cell(row=i, column=5).value
            data.expected = self.sheet.cell(row=i, column=6).value
            data.sql = self.sheet.cell(row=i, column=9).value
            big_data.append(data)
        self.workbook.close()

        return big_data

    # def write_data(self,row_num,column_num,new_data):
    #     self.sheet.cell(row_num,column_num).value = new_data
    #     self.wb.save(self.file)
    #     self.wb.close()

    def write_data(self,row,actual,result):#往指定的第7,8列，写入actual和result
        self.sheet.cell(row,7).value = actual
        self.sheet.cell(row,8).value = result
        self.workbook.save(self.filename)
        self.workbook.close()



if __name__ == '__main__':
    from common import contants
    r = DoExcel(contants.case_file,'login').read_data()
    print(r)
    DoExcel(contants.case_file, 'login').write_data(r[4].case_id+1,"hhh","123")

